# feature_family_analysis.py
# ==============================================================================
# Post-processing: map a cluster's stability features to concept families and
# rank the DRIVERS by magnitude. Reads EXISTING feature_stability_*.csv outputs
# (no re-run needed) + Annex Part B for the name -> family mapping.
#
# Answers: "which feature FAMILIES define this cluster, and how strongly?"
# e.g. drivers = Absorption (mean|z|=1.8, 6 feats), Aggression (1.4, 4 feats)...
#
# EXTERNAL PREREQUISITE (not committed to this repo):
#   Annex_PartB_Features.xlsx — the feature dictionary workbook, sheet
#   "PartB_Features", with columns concept_family / bare_name / base_concept.
#   It is maintained outside the codebase and must be supplied via --annexb.
#   If the path does not exist the script aborts with a clear message rather
#   than raising an opaque openpyxl error.
#
#   NOTE: the committed results/selection/feature_keep.csv carries an
#   equivalent column -> family mapping (it is the same source used by
#   cluster_engine._load_family_map) and can substitute for the workbook if
#   the Annex is unavailable.
#
# USAGE:
#   python feature_family_analysis.py \
#       --stability feature_stability_btc_5s_15bps_pca600_k6.csv \
#       --annexb Annex_PartB_Features.xlsx \
#       --out family_drivers_btc_5s_15bps.csv
# ==============================================================================
from __future__ import annotations
import argparse
import sys
from pathlib import Path
import pandas as pd
import numpy as np


def load_family_maps(annexb_path: str):
    """Return (exact_name->family, base_concept->family sorted by length desc)."""
    import openpyxl
    wb = openpyxl.load_workbook(annexb_path, read_only=True, data_only=True)
    ws = wb["PartB_Features"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: j for j, h in enumerate(hdr)}
    exact, base = {}, {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        fam = r[idx["concept_family"]]
        if not fam:
            continue
        bn = r[idx["bare_name"]]
        bc = r[idx["base_concept"]]
        if bn:
            exact[str(bn)] = fam
        if bc:
            base.setdefault(str(bc), fam)
    # longest base_concept first so the most specific prefix wins
    base_sorted = sorted(base.items(), key=lambda kv: -len(kv[0]))
    return exact, base_sorted


def tag_family(name: str, exact: dict, base_sorted: list) -> str:
    if name in exact:
        return exact[name]
    for bc, fam in base_sorted:
        if name == bc or name.startswith(bc + "_"):
            return fam
    return "UNMAPPED"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--stability", required=True,
                    help="feature_stability_*.csv produced by cluster_validation")
    ap.add_argument("--annexb", required=True, help="Annex_PartB_Features.xlsx")
    ap.add_argument("--out", default="family_drivers.csv")
    ap.add_argument("--zcol", default="mean_z",
                    help="z column to rank by (mean_z or full_z)")
    args = ap.parse_args()

    # The Annex workbook is an external, uncommitted prerequisite (see module
    # docstring). Fail fast with an actionable message instead of an opaque
    # openpyxl FileNotFoundError.
    if not Path(args.annexb).is_file():
        sys.exit(
            f"[feature_family_analysis] Annex workbook not found: {args.annexb}\n"
            f"  This file is an external prerequisite and is not committed to the repo.\n"
            f"  Supply it via --annexb, or substitute the committed\n"
            f"  results/selection/feature_keep.csv (column -> family) mapping."
        )

    df = pd.read_csv(args.stability)
    if args.zcol not in df.columns:
        # fall back gracefully
        args.zcol = "full_z" if "full_z" in df.columns else df.columns[1]
    exact, base_sorted = load_family_maps(args.annexb)

    df["family"] = df["feature"].astype(str).map(
        lambda n: tag_family(n, exact, base_sorted))
    df["abs_z"] = df[args.zcol].abs()
    df["direction"] = np.sign(df[args.zcol]).map({1: "+", -1: "-", 0: "0"})

    # per-feature, sorted by magnitude
    feat = df.sort_values("abs_z", ascending=False)[
        ["feature", "family", args.zcol, "abs_z", "direction"]]

    # per-family aggregation = the "drivers"
    fam = (df.groupby("family")
             .agg(n_features=("feature", "count"),
                  sum_abs_z=("abs_z", "sum"),
                  mean_abs_z=("abs_z", "mean"),
                  max_abs_z=("abs_z", "max"))
             .sort_values("sum_abs_z", ascending=False)
             .round(3))

    unmapped = df.loc[df["family"] == "UNMAPPED", "feature"].tolist()

    print(f"\n=== FAMILY DRIVERS  ({Path(args.stability).name}, ranked by {args.zcol}) ===")
    print(fam.to_string())
    print(f"\n=== TOP FEATURES ===")
    print(feat.head(20).to_string(index=False))
    if unmapped:
        print(f"\n[!] {len(unmapped)} unmapped feature(s): {unmapped[:10]}"
              f"{' ...' if len(unmapped) > 10 else ''}")
        print("    -> extend the name normaliser or add to Annex B before trusting families.")

    fam.to_csv(args.out)
    print(f"\nSaved family-level drivers -> {args.out}")


if __name__ == "__main__":
    main()